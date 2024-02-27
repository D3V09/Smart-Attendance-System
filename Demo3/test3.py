import datetime
import cv2
from pyzbar.pyzbar import decode
import tkinter as tk
from PIL import Image, ImageTk
import openpyxl
import keyboard
import threading

# def store_data_in_excel(data, row):
#     excel_file_path = u'C:/Users/darsh/Documents/PayasTechnology/GivenProjects/Attendance/Demo3/Attendance.xlsx'
#     workbook = openpyxl.load_workbook(excel_file_path)
#     sheet = workbook.active
#     sheet.cell(row=row, column=1, value=data)
#     workbook.save(excel_file_path)

from openpyxl import Workbook, load_workbook

def save_to_excel(data):
    file_path = u'C:/Users/darsh/Documents/PayasTechnology/GivenProjects/Attendance/Demo3/Attendance.xlsx'
    try:
        wb = load_workbook(file_path)
        ws = wb.active
    except FileNotFoundError:
        wb = Workbook()
        ws = wb.active
        ws.append(["QR Data"])  # Add headers if creating a new workbook
    ws.append([data])
    wb.save(file_path)


def scan_qr_code(frame_label):
    cap = cv2.VideoCapture(0)
    current_row = 1

    last_data = None

    while True:
        ret, frame = cap.read()

        frame_rgb = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)

        decoded_objects = decode(frame_rgb)

        if decoded_objects:
            for obj in decoded_objects:
                qr_data = obj.data.decode('utf-8')+" "+str(datetime.datetime.now())

                check = qr_data.split()[:2]
                if check != last_data:
                    last_data = qr_data.split()[:2]
                    save_to_excel(qr_data)
                    # frame_label.config("Attendance Marked")
                    current_row += 1
                else:
                    print("Duplicate data detected. Skipping...")

        frame_pil = Image.fromarray(frame_rgb)
        frame_tk = ImageTk.PhotoImage(image=frame_pil)
        frame_label.config(image=frame_tk)
        frame_label.image = frame_tk

def start_scanning(frame_label):
    scanning_thread = threading.Thread(target=scan_qr_code, args=(frame_label,))
    scanning_thread.daemon = True
    scanning_thread.start()

def main():
    root = tk.Tk()
    root.title("QR Code Scanner")

    screen_width = root.winfo_screenwidth()
    screen_height = root.winfo_screenheight()
    root.geometry(f"{screen_width}x{screen_height}+0+0")

    frame_label = tk.Label(root)
    frame_label.pack(padx=10, pady=10)
    root.update_idletasks()
    frame_label.place(relx=0.5, rely=0.5, anchor=tk.CENTER)

    start_scanning(frame_label)
    # root.bind("<q>", frame_label.destroy())
    root.mainloop()


if __name__ == "__main__":
    main()
